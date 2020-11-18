import datetime
import os

import xlsxwriter
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font

from com.sca.hem4.writer.Writer import Writer

class ExcelWriter(Writer):

    def __init__(self, model, plot_df):
        Writer.__init__(self)

        self.model = model
        self.plot_df = plot_df

    def writeWithTimestamp(self):
        if os.path.exists(self.filename):
            os.remove(self.filename)

        # Create a blank workbook
        workbook = xlsxwriter.Workbook(self.filename, {'constant_memory': True})
        workbook.add_worksheet()
        workbook.close()

        # Note: we are removing the timestamp and extra lines at the request of Mark from EPA.
        # These extra lines at the top make pandastable work a little funny!
        #self.insertTimestamp()
        self.appendHeaderAtLocation(self.getHeader(), startingrow=0)
        for data in self.generateOutputs():
            if data is not None:
                self.appendToFile(data)
                self.analyze(data)

    def insertTimestamp(self):
        now = datetime.datetime.now()
        timestamp = now.strftime('Created on %A, %B %d, %Y @ %I:%M %p')
        timestamp_df = pd.DataFrame([timestamp])
        self.appendToFileAtLocation(dataframe=timestamp_df)

    def appendToFile(self, dataframe):
        data = dataframe.values
        book = load_workbook(self.filename)
        writer = pd.ExcelWriter(self.filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        ws = writer.book["Sheet1"]
        startrow = ws.max_row
        for row in range(0, data.shape[0]):
            for col in range(0, data.shape[1]):
                value = data[row][col]
                truncated = float('{:6.12}'.format(value)) if isinstance(value, float) else value
                ws.cell(row=startrow + row+1, column=col+1).value = truncated

        writer.save()

    def appendToFileAtLocation(self, dataframe, startingrow=None, startingcol=None):
        data = dataframe.values
        book = load_workbook(self.filename)
        writer = pd.ExcelWriter(self.filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        ws = writer.book["Sheet1"]
        startrow = ws.max_row if startingrow is None else startingrow
        startcol = 0 if startingcol is None else startingcol
        for row in range(0, data.shape[0]):
            for col in range(0, data.shape[1]):
                value = data[row][col]
                truncated = float('{:6.12}'.format(value)) if isinstance(value, float) else value
                ws.cell(row=startrow+row+1, column=startcol+col+1).value = truncated

        writer.save()
        book.close()

    def appendHeaderAtLocation(self, headers, startingrow=None, startingcol=None):
        book = load_workbook(self.filename)
        writer = pd.ExcelWriter(self.filename, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        ws = writer.book["Sheet1"]

        ft = Font(bold=True)
        startrow = ws.max_row if startingrow is None else startingrow
        startcol = 0 if startingcol is None else startingcol
        for i in range(0, len(headers)):
            ws.cell(row=startrow+1, column=startcol+i+1).font = ft
            ws.cell(row=startrow+1, column=startcol+i+1).value = headers[i]

        writer.save()
        book.close()

    def writeHeader(self):
        """
         Write the header (column names) in the given Excel worksheet.
        """
        workbook = xlsxwriter.Workbook(self.filename, {'constant_memory': True})
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True})

        headers = self.getHeader()
        for i in range(0, len(headers)):
            worksheet.write(0, i, headers[i], bold)

        workbook.close()

    def analyze(self, data):
        pass